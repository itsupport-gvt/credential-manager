#!/usr/bin/env python3
"""
Credential Manager Excel Workbook Generator
============================================
Creates a comprehensive, SharePoint-compatible Excel workbook for managing
credentials across multiple tenants/clients.

Style: black header rows with white font, plain white data rows, minimal borders.

Sheets produced:
  _README              - Instructions and legend
  Master_Credentials   - Single-table datastore (all credentials)
  Change_Log           - Full audit trail
  Tenants              - Client/tenant reference
  Categories           - Category & subcategory reference
  Internal_Users       - Staff who manage credentials
  Dashboard            - Live summary (COUNTIFS formulas)
  VIEW_M365_Admin      - Live FILTER view: Microsoft 365
  VIEW_Azure           - Live FILTER view: Azure
  VIEW_Domains_FTP     - Live FILTER view: Domains / Hosting / FTP
  VIEW_Social_Media    - Live FILTER view: Social Media
  VIEW_Marketing       - Live FILTER view: Marketing Tools
  VIEW_Email           - Live FILTER view: Email Systems
  VIEW_Expiring_90d    - Live FILTER view: Expiring within 90 days
"""

import os
from datetime import datetime, date

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

# ──────────────────────────────────────────────────────────────────────────────
# OUTPUT
# ──────────────────────────────────────────────────────────────────────────────
OUTPUT_DIR  = r"c:\Dev\Credential_manager"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Credential_Manager.xlsx")
TODAY       = date.today()
NOW_STR     = datetime.now().strftime("%Y-%m-%d")

# ──────────────────────────────────────────────────────────────────────────────
# STYLE CONSTANTS
# ──────────────────────────────────────────────────────────────────────────────
BLACK   = "111111"   # pure dark for headers / title
DGREY   = "333333"   # very dark grey for sub-headers
WHITE   = "FFFFFF"
LTGREY  = "EAEAEA"   # minimal border colour
TXTCLR  = "111111"   # body text

FONT_NAME = "Segoe UI"

# ──────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ──────────────────────────────────────────────────────────────────────────────
def fl(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def fn(color=TXTCLR, size=9, bold=False, italic=False):
    return Font(name=FONT_NAME, color=color, size=size, bold=bold, italic=italic)

def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color=LTGREY)
    return Border(left=s, right=s, top=s, bottom=s)

def no_border():
    return Border()

# ── Shared cell writers ────────────────────────────────────────────────────────

def write_title(ws, row, ncols, text, height=32):
    """Full-width black title row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.value  = text
    cell.fill   = fl(BLACK)
    cell.font   = Font(name=FONT_NAME, color=WHITE, size=12, bold=True)
    cell.alignment = al("left", "center")
    ws.row_dimensions[row].height = height

def write_group_header(ws, row, c1, c2, text):
    """Dark-grey group label spanning columns c1..c2."""
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value  = text
    cell.fill   = fl(DGREY)
    cell.font   = fn(WHITE, 8, bold=True)
    cell.alignment = al("center", "center")
    ws.row_dimensions[row].height = 16

def write_header_cell(ws, row, col, text):
    """Single column-header cell: black fill, white bold text."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill      = fl(BLACK)
    cell.font      = fn(WHITE, 9, bold=True)
    cell.alignment = al("center", "center", wrap=True)
    cell.border    = thin_border()
    return cell

def write_data_cell(ws, row, col, value, num_fmt=None):
    """Plain white data cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = fn(TXTCLR, 9)
    cell.alignment = al("left", "center", wrap=True)
    cell.border    = thin_border()
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def write_info_cell(ws, row, ncols, text):
    """Full-width italic note row (light grey)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.value      = text
    cell.fill       = fl("F9F9F9")
    cell.font       = fn("555555", 9, italic=True)
    cell.alignment  = al("left", "center", wrap=True)
    ws.row_dimensions[row].height = 20

def add_table(ws, name, ref, style="TableStyleLight1"):
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)

def dv_list(formula1, sqref):
    dv = DataValidation(
        type="list", formula1=formula1,
        allow_blank=True, showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Value",
        error="Please choose a value from the dropdown list.",
    )
    dv.sqref = sqref
    return dv

# ──────────────────────────────────────────────────────────────────────────────
# MASTER TABLE COLUMN DEFINITIONS  (header_text, col_width)
# ──────────────────────────────────────────────────────────────────────────────
MASTER_COLS = [
    # CORE IDENTITY
    ("CredentialID",          16),
    ("Tenant_Code",           11),
    ("Tenant_Name",           24),
    ("Category",              22),
    ("Subcategory",           22),
    ("Service_Name",          24),
    ("Service_URL",           36),
    ("Environment",           14),
    ("Status",                14),
    ("Priority",              11),
    # AUTHENTICATION
    ("Username_Email",        30),
    ("Password",              24),
    ("Recovery_Email",        30),
    ("Recovery_Phone",        17),
    ("MFA_Enabled",           11),
    ("MFA_Type",              20),
    ("MFA_App_Name",          18),
    ("Backup_Codes_Location", 28),
    ("Security_Notes",        28),
    # ACCOUNT DETAILS
    ("Account_Display_Name",  26),
    ("Account_ID",            24),
    ("License_Type",          20),
    ("Plan_Tier",             16),
    ("Subscription_Start",    15),
    ("Subscription_End",      15),
    ("Auto_Renewal",          12),
    ("Monthly_Cost",          13),
    ("Billing_Cycle",         14),
    ("Billing_Email",         30),
    ("Payment_Reference",     24),
    # TECHNICAL / API
    ("Access_Level",          16),
    ("Linked_CredentialID",   18),
    ("API_Key",               36),
    ("API_Secret",            36),
    ("Client_ID",             36),
    ("Client_Secret",         36),
    ("Tenant_ID_App",         36),
    ("Subscription_ID_Azure", 36),
    ("Server_Hostname",       28),
    ("Port",                   8),
    ("Protocol",              13),
    ("Database_Name",         18),
    # OWNERSHIP & TRACKING
    ("Managed_By",            22),
    ("Managed_By_Email",      30),
    ("Created_By",            22),
    ("Created_Date",          14),
    ("Last_Updated_By",       22),
    ("Last_Updated_Date",     14),
    ("Last_Verified_Date",    14),
    ("Last_Password_Changed", 14),
    ("Password_Expiry_Date",  14),
    ("Next_Review_Date",      14),
    ("Tags",                  26),
    ("Notes",                 42),
    ("Record_Status",         14),
]

NCOLS_MASTER = len(MASTER_COLS)

MASTER_GROUPS = [
    ("CORE IDENTITY",        "CredentialID",        "Priority"),
    ("AUTHENTICATION",       "Username_Email",       "Security_Notes"),
    ("ACCOUNT DETAILS",      "Account_Display_Name", "Payment_Reference"),
    ("TECHNICAL / API",      "Access_Level",         "Database_Name"),
    ("OWNERSHIP & TRACKING", "Managed_By",           "Record_Status"),
]

def col_of(name):
    for i, (h, _) in enumerate(MASTER_COLS, 1):
        if h == name:
            return i
    raise ValueError(f"Column {name!r} not found")

# ──────────────────────────────────────────────────────────────────────────────
# OTHER TABLE COLUMNS
# ──────────────────────────────────────────────────────────────────────────────
CHANGELOG_COLS = [
    ("LogID",            14),
    ("Timestamp",        18),
    ("CredentialID",     16),
    ("Tenant_Code",      11),
    ("Tenant_Name",      24),
    ("Service_Name",     24),
    ("Action",           20),
    ("Field_Changed",    22),
    ("Old_Value_Masked", 30),
    ("New_Value_Masked", 30),
    ("Changed_By",       22),
    ("Changed_By_Email", 30),
    ("Reason",           30),
    ("Notes",            40),
]
NCOLS_CHANGELOG = len(CHANGELOG_COLS)

TENANT_COLS = [
    ("TenantID",        12),
    ("Tenant_Code",     12),
    ("Tenant_Name",     28),
    ("Industry",        22),
    ("Primary_Contact", 24),
    ("Contact_Email",   30),
    ("Contact_Phone",   18),
    ("Account_Manager", 24),
    ("Contract_Start",  15),
    ("Contract_End",    15),
    ("Status",          14),
    ("Notes",           40),
]
NCOLS_TENANTS = len(TENANT_COLS)

CATEGORY_COLS = [
    ("CategoryID",     12),
    ("Category_Name",  28),
    ("Category_Code",  14),
    ("Description",    40),
    ("Subcategories",  65),
]
NCOLS_CATEGORIES = len(CATEGORY_COLS)

USER_COLS = [
    ("UserID",       12),
    ("Full_Name",    24),
    ("Email",        30),
    ("Role",         22),
    ("Department",   20),
    ("Access_Level", 16),
    ("Status",       14),
    ("Notes",        36),
]
NCOLS_USERS = len(USER_COLS)

# ──────────────────────────────────────────────────────────────────────────────
# VALIDATION LISTS
# ──────────────────────────────────────────────────────────────────────────────
STATUSES      = ["Active", "Inactive", "Expired", "Compromised", "Archived"]
PRIORITIES    = ["Critical", "High", "Medium", "Low"]
ENVIRONMENTS  = ["Production", "Staging", "Development", "Test"]
MFA_ENABLED   = ["Yes", "No"]
MFA_TYPES     = ["Authenticator App", "SMS", "Email", "Hardware Key", "None"]
ACCESS_LEVELS = ["Owner", "Admin", "Editor", "Viewer", "API Only", "Read-Only"]
PROTOCOLS     = ["HTTPS", "SFTP", "FTP", "FTPS", "SSH", "RDP", "MySQL", "SMTP", "IMAP", "POP3"]
BILLING_CYCLES= ["Monthly", "Annual", "Bi-Annual", "One-Time", "Free"]
AUTO_RENEWAL  = ["Yes", "No"]
RECORD_STATUS = ["Active", "Archived"]
LOG_ACTIONS   = [
    "Created", "Updated", "Password Changed", "Accessed", "Reviewed",
    "Archived", "Restored", "Deleted", "Compromised", "Shared", "Revoked",
]
TENANT_STATUSES = ["Active", "Inactive", "Prospect", "Churned"]
USER_ROLES    = ["Admin", "Manager", "Technician", "Viewer", "Auditor"]
USER_ACCESS   = ["Full Access", "Manage Only", "View Only"]
USER_STATUSES = ["Active", "Inactive", "On Leave"]

# ──────────────────────────────────────────────────────────────────────────────
# REFERENCE DATA
# ──────────────────────────────────────────────────────────────────────────────
CATEGORIES_DATA = [
    ("CAT-M365", "Microsoft 365", "M365", "Microsoft 365 tenant and service accounts",
     "Global Admin; Exchange Admin; SharePoint Admin; Teams Admin; User Account; Mailbox Access (OTP/SSO); Service Account"),
    ("CAT-CLOUD", "Cloud Infrastructure", "CLOUD", "Amazon Web Services, Microsoft Azure, DigitalOcean and other cloud providers",
     "AWS Root; AWS IAM User; Azure Portal; Azure Sub Owner; DigitalOcean; Google Cloud Portal; Service Principal"),
    ("CAT-DNS", "Domain & DNS", "DNS", "Domain registrars, DNS hosting and SSL certificates",
     "Domain Registrar; DNS Provider; SSL/TLS Certificate; CDN; Domain Forwarding"),
    ("CAT-HOST", "Web Hosting & Panels", "HOST", "Server control panels, cPanel, WHM, Plesk, and hosting portal access",
     "cPanel; Plesk; WHM; Web Host Portal; Virtual Private Server (VPS); Managed Hosting"),
    ("CAT-CMS", "CMS & Websites", "CMS", "Content Management Systems (WordPress, Django, CMS portals) and site admin panels",
     "WordPress Admin; Django Admin; Shopify Admin; Webflow; Squarespace; Custom CMS"),
    ("CAT-GOOG", "Google Ecosystem", "GOOG", "Google Search Console, Google Business Profile, Analytics and Google account credentials",
     "Google Search Console; Google Business Profile (GMB); Google Analytics; Google Tag Manager; Workspace Admin"),
    ("CAT-DB", "Databases", "DB", "Database administration, SQL Server, Azure SQL, MySQL, and PostgreSQL logins",
     "Azure SQL; MySQL/MariaDB; PostgreSQL; Microsoft SQL Server; MongoDB; Redis; Database Administrator"),
    ("CAT-FTP", "FTP & File Sharing", "FTP", "File transfer protocol (FTP, SFTP) credentials and cloud/network storage vaults",
     "SFTP; FTP; FTPS; S3 Bucket; WebDAV; Network File Share"),
    ("CAT-ERP", "ERP & Business Portals", "ERP", "Zoho, Odoo, task workflow portals, and learning platforms",
     "Zoho Admin; Zoho User; Odoo Admin; Odoo User; ERP Portal; Task Assignment; Workflow Portal"),
    ("CAT-MKT", "Marketing & Creative", "MKT", "Marketing tools, graphic design, and creative suite applications",
     "Adobe Creative Cloud; Graphic Design Tool; Email Marketing (Mailchimp); SEO/SEM Tool; Marketing Analytics; Canva"),
    ("CAT-SOC", "Social Media", "SOC", "Brand social media handles and ad manager accounts",
     "Facebook/Meta; Instagram; Twitter/X; LinkedIn; YouTube; TikTok; Pinterest"),
    ("CAT-AI", "AI & Chatbots", "AI", "Artificial Intelligence systems, LLM tools, and AI developer portals",
     "ChatGPT (OpenAI); Claude (Anthropic); Gemini (Google); Copilot; AI API Key; Custom Chatbot"),
    ("CAT-HR", "HR & Recruitment", "HR", "HR management systems (HRMS), job portals (Naukri, Indeed), and payroll portal access",
     "HRMS Portal; Recruitment Portal (Naukri/Indeed); Payroll System; Attendance/Leave System"),
    ("CAT-FIN", "Finance & Accounting", "FIN", "Accounting softwares (Zoho Admin, Odoo Admin, users, etc.) and invoicing platforms",
     "Zoho Books/Finance; Odoo Accounting; Accounting Software; Payment Gateway; Billing/Invoice Portal"),
    ("CAT-WORK", "Workflow & Productivity", "WORK", "Team collaboration, task tracking, and learning portals",
     "Jira; Monday.com; Asana; Trello; ClickUp; Slack; Microsoft Teams; Learning Portal (Udemy)"),
    ("CAT-NET", "Network & CCTV", "NET", "Router/switch logins, internal WiFi, and CCTV/NVR surveillance cameras",
     "WiFi Router/Access Point; Firewall/Switch; CCTV Camera/NVR; Network Attached Storage (NAS); ISP Admin Portal"),
    ("CAT-TEL", "Telecom & Utility", "TEL", "Network service providers, postpaid sim portals, and broadband logins",
     "Postpaid SIM Portal; WiFi Service Provider; Telecom Login; Broadband Portal"),
    ("CAT-OTH", "Other", "OTH", "Miscellaneous credentials not covered in above categories",
     "General; Utility; Hardware/IoT; Security Vault"),
]

TENANT_DATA = [
    ("T-001", "GBP", "Gravity Business Partners",  "IT Services",     "Dev Mukherjee",    "info@gravity-bp.com",      "+971 XX XXX XXXX", "Dev Mukherjee",    "2020-01-01", "",           "Active",   "Internal - own company"),
    ("T-002", "TAS", "TechAlpha Solutions",         "Technology",      "Ahmed Al Sayed",   "admin@techalpha.ae",       "+971 50 XXX XXXX", "Dev Mukherjee",    "2023-03-15", "2026-03-14", "Active",   ""),
    ("T-003", "BKR", "Bakery King LLC",             "Food & Beverage", "Sara Mahmoud",     "sara@bakeryking.com",      "+971 55 XXX XXXX", "Dev Mukherjee",    "2022-07-01", "2025-06-30", "Active",   "Multiple outlets"),
    ("T-004", "MED", "MediCare Clinics Group",      "Healthcare",      "Dr. Ravi Sharma",  "admin@medicareclinics.ae", "+971 4 XXX XXXX",  "Dev Mukherjee",    "2024-01-15", "2026-01-14", "Active",   "Sensitive - restrict access"),
    ("T-005", "REA", "RealEstate Pro PJSC",         "Real Estate",     "Khalid Al Rashid", "khalid@realestatepro.ae",  "+971 52 XXX XXXX", "Dev Mukherjee",    "2023-10-01", "2025-09-30", "Active",   ""),
]

USER_DATA = [
    ("USR-001", "Dev Mukherjee",  "info@gravity-bp.com",    "Admin",      "IT",         "Full Access",  "Active", "Primary admin"),
    ("USR-002", "Support Team",   "support@gravity-bp.com", "Technician", "IT",         "Manage Only",  "Active", ""),
    ("USR-003", "Audit User",     "audit@gravity-bp.com",   "Auditor",    "Compliance", "View Only",    "Active", "Read-only audit access"),
]

# ──────────────────────────────────────────────────────────────────────────────
# SAMPLE CREDENTIALS
# ──────────────────────────────────────────────────────────────────────────────
def make_cred(
    cid, tc, tn, cat, sub, svc, url, env, status, pri,
    user, pw, rec_email, rec_ph, mfa, mfa_type, mfa_app, bkp, sec_notes,
    acc_name, acc_id, lic, plan, sub_start, sub_end, renew, cost, bcycle, bill_email, pay_ref,
    acc_lvl, linked, api_key, api_sec, cli_id, cli_sec, tid, sub_id_az, host, port, proto, db,
    mgr, mgr_email, cby, cdate, uby, udate, vdate, pwchg, pwexp, review, tags, notes, rec_st,
):
    keys = [h for h, _ in MASTER_COLS]
    vals = [
        cid, tc, tn, cat, sub, svc, url, env, status, pri,
        user, pw, rec_email, rec_ph, mfa, mfa_type, mfa_app, bkp, sec_notes,
        acc_name, acc_id, lic, plan, sub_start, sub_end, renew, cost, bcycle, bill_email, pay_ref,
        acc_lvl, linked, api_key, api_sec, cli_id, cli_sec, tid, sub_id_az, host, port, proto, db,
        mgr, mgr_email, cby, cdate, uby, udate, vdate, pwchg, pwexp, review, tags, notes, rec_st,
    ]
    return dict(zip(keys, vals))

SAMPLE_CREDENTIALS = [
    make_cred(
        "CRED-2024-0001","GBP","Gravity Business Partners",
        "Microsoft 365","Global Admin","Microsoft 365 - GBP Tenant",
        "https://admin.microsoft.com","Production","Active","Critical",
        "admin@gravity-bp.com","P@ssw0rd!GBP2024","info@gravity-bp.com","+971 XX XXX XXXX",
        "Yes","Authenticator App","Microsoft Authenticator","1Password Vault - GBP","Global admin - break-glass account",
        "Gravity BP Global Admin","xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx",
        "Microsoft 365 Business Premium","Business Premium",
        date(2020,1,1),date(2025,12,31),"Yes",22.00,"Monthly","billing@gravity-bp.com","Visa ending 4242",
        "Owner","","","","","","xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx","",
        "","","HTTPS","",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2020,1,1),"Dev Mukherjee",date(2024,6,1),
        date(2024,12,1),date(2024,6,1),date(2025,6,1),date(2025,6,1),
        "M365; Admin; Critical","Primary M365 global admin. Store in 1Password.","Active",
    ),
    make_cred(
        "CRED-2024-0002","GBP","Gravity Business Partners",
        "Cloud Infrastructure","Azure Sub Owner","Azure - GBP Production Subscription",
        "https://portal.azure.com","Production","Active","Critical",
        "admin@gravity-bp.com","(uses M365 SSO)","","",
        "Yes","Authenticator App","Microsoft Authenticator","1Password Vault - GBP","SSO via Microsoft 365 account",
        "GBP Production","xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx",
        "Pay-As-You-Go","Pay-As-You-Go",
        date(2020,6,1),"","Yes",0,"Monthly","billing@gravity-bp.com","Visa ending 4242",
        "Owner","CRED-2024-0001","","","","",
        "xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx","xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx",
        "portal.azure.com","443","HTTPS","",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2020,6,1),"Dev Mukherjee",date(2024,6,1),
        date(2024,12,1),"","",date(2025,6,1),
        "Azure; Cloud; Critical","Linked to M365 Global Admin (CRED-2024-0001).","Active",
    ),
    make_cred(
        "CRED-2024-0003","GBP","Gravity Business Partners",
        "Social Media","Facebook/Meta","Facebook - Gravity Business Partners Page",
        "https://www.facebook.com/gravitybp","Production","Active","High",
        "social@gravity-bp.com","Fb@GBP2024!","info@gravity-bp.com","+971 XX XXX XXXX",
        "Yes","SMS","","Recovery codes in 1Password","",
        "Gravity Business Partners","xxxxxxxxxxxxxxxxxx",
        "","Business Suite",
        "","","No",0,"Free","","",
        "Admin","","","","","","","",
        "","","HTTPS","",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2021,3,15),"Dev Mukherjee",date(2024,5,1),
        date(2024,11,1),date(2024,5,1),"",date(2025,5,1),
        "Social; Facebook; Marketing","Business Manager also linked. Ads account separate.","Active",
    ),
    make_cred(
        "CRED-2024-0004","TAS","TechAlpha Solutions",
        "Domain & DNS","Domain Registrar","Namecheap - techalpha.ae",
        "https://www.namecheap.com","Production","Active","Critical",
        "admin@techalpha.ae","NC!T@s2024Secure","it@techalpha.ae","+971 50 XXX XXXX",
        "Yes","Authenticator App","Google Authenticator","Printed backup in client file","",
        "TechAlpha Domains","",
        "","",
        date(2023,3,15),date(2026,3,14),"Yes",12.98,"Annual","billing@techalpha.ae","Card ending 1234",
        "Admin","","","","","","","",
        "","","HTTPS","",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2023,3,15),"Dev Mukherjee",date(2024,3,15),
        date(2024,9,15),date(2024,3,15),"",date(2025,3,15),
        "Domain; DNS; Critical; TAS","Domain renewal due 2026-03-14. Auto-renewal ON.","Active",
    ),
    make_cred(
        "CRED-2024-0005","TAS","TechAlpha Solutions",
        "Web Hosting & Panels","cPanel","SiteGround cPanel - techalpha.ae",
        "https://my.siteground.com","Production","Active","High",
        "admin@techalpha.ae","SG!cP@nel2024","it@techalpha.ae","",
        "Yes","Authenticator App","Google Authenticator","1Password - TAS","",
        "TechAlpha SiteGround","",
        "GrowBig Plan","GrowBig",
        date(2023,3,15),date(2025,3,14),"Yes",22.99,"Annual","billing@techalpha.ae","Card ending 1234",
        "Admin","CRED-2024-0004","","","","","","",
        "sg-srv12.siteground.com","443","HTTPS","",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2023,3,15),"Dev Mukherjee",date(2024,3,15),
        date(2024,9,15),date(2024,3,15),date(2025,3,14),date(2025,3,1),
        "Hosting; cPanel; TAS","Hosting expires 2025-03-14 - review 3 months before.","Active",
    ),
    make_cred(
        "CRED-2024-0006","TAS","TechAlpha Solutions",
        "FTP & File Sharing","SFTP","SFTP - techalpha.ae Production Server",
        "","Production","Active","High",
        "sftp_tas","SFTP!TAS2024@#","","",
        "No","None","","","Key-based auth also available - key stored in 1Password",
        "","",
        "","",
        "","","",0,"","","",
        "Admin","CRED-2024-0005","","","","","","",
        "sg-srv12.siteground.com","22","SFTP","public_html",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2023,3,15),"Dev Mukherjee",date(2024,6,1),
        date(2024,12,1),date(2024,6,1),"",date(2025,6,1),
        "FTP; SFTP; Server; TAS","Linked to cPanel (CRED-2024-0005).","Active",
    ),
    make_cred(
        "CRED-2024-0007","BKR","Bakery King LLC",
        "Google Ecosystem","Workspace Admin","Google Workspace - bakeryking.com",
        "https://admin.google.com","Production","Active","Critical",
        "admin@bakeryking.com","BK@GW2024Admin!","sara@bakeryking.com","+971 55 XXX XXXX",
        "Yes","Authenticator App","Google Authenticator","Stored with client","",
        "Bakery King Workspace Admin","",
        "Google Workspace Business Starter","Business Starter",
        date(2022,7,1),date(2025,6,30),"Yes",6.00,"Monthly","accounts@bakeryking.com","Card ending 5678",
        "Admin","","","","","","","",
        "","","HTTPS","",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2022,7,1),"Dev Mukherjee",date(2024,7,1),
        date(2024,12,1),date(2024,7,1),"",date(2025,7,1),
        "Email; Google; Admin; BKR","10 user licences. Billing via client card.","Active",
    ),
    make_cred(
        "CRED-2024-0008","BKR","Bakery King LLC",
        "Social Media","Instagram","Instagram - @bakerykinguae",
        "https://www.instagram.com/bakerykinguae","Production","Active","High",
        "social@bakeryking.com","IG!BKR2024#","sara@bakeryking.com","+971 55 XXX XXXX",
        "Yes","Authenticator App","Google Authenticator","With client","",
        "Bakery King UAE","xxxxxxxxxxxxxxxxxxx",
        "","Business",
        "","","No",0,"Free","","",
        "Admin","","","","","","","",
        "","","HTTPS","",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2022,8,1),"Dev Mukherjee",date(2024,5,1),
        date(2024,11,1),date(2024,5,1),"",date(2025,5,1),
        "Social; Instagram; BKR","Linked to Facebook Business Manager.","Active",
    ),
    make_cred(
        "CRED-2024-0009","MED","MediCare Clinics Group",
        "Microsoft 365","Global Admin","Microsoft 365 - MediCare Tenant",
        "https://admin.microsoft.com","Production","Active","Critical",
        "itsupport@medicareclinics.ae","MC!M365Adm2024","admin@medicareclinics.ae","+971 4 XXX XXXX",
        "Yes","Authenticator App","Microsoft Authenticator","1Password - MED",
        "HIPAA-sensitive - restrict access to senior staff only",
        "MediCare Global Admin","xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx",
        "Microsoft 365 Business Premium","Business Premium",
        date(2024,1,15),date(2026,1,14),"Yes",22.00,"Monthly","it@medicareclinics.ae","Card ending 9012",
        "Owner","","","","","","xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx","",
        "","","HTTPS","",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2024,1,15),"Dev Mukherjee",date(2024,6,1),
        date(2024,12,1),date(2024,6,1),date(2025,6,1),date(2025,6,1),
        "M365; Admin; Critical; Healthcare","HIPAA environment - log all access. Break-glass only.","Active",
    ),
    make_cred(
        "CRED-2024-0010","REA","RealEstate Pro PJSC",
        "Marketing & Creative","Email Marketing (Mailchimp)","Mailchimp - RealEstate Pro Newsletter",
        "https://mailchimp.com","Production","Active","Medium",
        "marketing@realestatepro.ae","MC!REA2024Mkt","khalid@realestatepro.ae","",
        "Yes","Email","","Recovery via account email","",
        "RealEstate Pro Mailchimp","",
        "Essentials","Essentials - 5000 contacts",
        date(2023,10,1),date(2025,9,30),"Yes",13.00,"Monthly","marketing@realestatepro.ae","Card ending 3456",
        "Admin","","","","","","","",
        "","","HTTPS","",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2023,10,1),"Dev Mukherjee",date(2024,4,1),
        date(2024,10,1),date(2024,4,1),date(2025,9,30),date(2025,9,1),
        "Marketing; Email; Mailchimp; REA","5000 contact limit - monitor subscriber count.","Active",
    ),
    make_cred(
        "CRED-2024-0011","REA","RealEstate Pro PJSC",
        "Marketing & Creative","SEO/SEM Tool","HubSpot CRM - RealEstate Pro",
        "https://app.hubspot.com","Production","Active","High",
        "crm@realestatepro.ae","HS!REA2024CRM","khalid@realestatepro.ae","",
        "Yes","Email","","Recovery via account email","",
        "RealEstate Pro HubSpot","",
        "Sales Hub Starter","Starter",
        date(2023,10,1),date(2025,9,30),"Yes",50.00,"Monthly","billing@realestatepro.ae","Card ending 3456",
        "Admin","","","","","","","",
        "","","HTTPS","",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2023,10,1),"Dev Mukherjee",date(2024,4,1),
        date(2024,10,1),date(2024,4,1),date(2025,9,30),date(2025,9,1),
        "Marketing; CRM; HubSpot; REA","2 seats included. API key stored separately.","Active",
    ),
    make_cred(
        "CRED-2024-0012","GBP","Gravity Business Partners",
        "Social Media","LinkedIn","LinkedIn - Gravity Business Partners Page",
        "https://www.linkedin.com/company/gravitybp","Production","Active","High",
        "info@gravity-bp.com","(uses personal M365 account)","","",
        "Yes","Authenticator App","Microsoft Authenticator","","Page admin via personal LinkedIn of Dev Mukherjee",
        "Gravity Business Partners","",
        "","Company Page",
        "","","No",0,"Free","","",
        "Admin","CRED-2024-0001","","","","","","",
        "","","HTTPS","",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2020,1,1),"Dev Mukherjee",date(2024,1,1),
        date(2024,10,1),"","",date(2025,1,1),
        "Social; LinkedIn; GBP","Managed via personal account. Add company admin if staff changes.","Active",
    ),
    make_cred(
        "CRED-2024-0013","TAS","TechAlpha Solutions",
        "CMS & Websites","WordPress Admin","WordPress - techalpha.ae",
        "https://techalpha.ae/wp-admin","Production","Active","High",
        "admin","WP!TAS2024@Admin","it@techalpha.ae","",
        "No","None","","","Wordfence active. Limit login attempts on.",
        "TechAlpha WP Admin","",
        "","",
        "","","",0,"Free","","",
        "Admin","CRED-2024-0005","","","","","","",
        "sg-srv12.siteground.com","443","HTTPS","wp_techalpha",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2023,3,15),"Dev Mukherjee",date(2024,6,1),
        date(2024,12,1),date(2024,6,1),"",date(2025,6,1),
        "CMS; WordPress; TAS","Linked to cPanel (CRED-2024-0005). DB on same server.","Active",
    ),
    make_cred(
        "CRED-2024-0014","MED","MediCare Clinics Group",
        "ERP & Business Portals","Zoho Admin","Zoho Mail - medicareclinics.ae",
        "https://mail.zoho.com/zm/","Production","Inactive","Medium",
        "postmaster@medicareclinics.ae","ZM!MED2023Old","admin@medicareclinics.ae","",
        "No","None","","","Migrated to M365 Exchange. Kept for archival only.",
        "MediCare Zoho Mail","",
        "Free Plan","Free",
        date(2020,1,1),date(2024,1,14),"No",0,"Free","","",
        "Admin","","","","","","","",
        "","","HTTPS","",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2020,1,1),"Dev Mukherjee",date(2024,1,15),
        date(2024,1,15),date(2024,1,15),date(2024,1,14),"",
        "Email; Zoho; MED; Inactive","INACTIVE - migrated to M365. Keep for archive email access.","Active",
    ),
    make_cred(
        "CRED-2024-0015","GBP","Gravity Business Partners",
        "Other","General","GitHub - GravityBP Org",
        "https://github.com/gravitybp","Production","Active","High",
        "dev@gravity-bp.com","(OAuth/SSO)","info@gravity-bp.com","",
        "Yes","Authenticator App","GitHub Authenticator","1Password - GBP",
        "Personal access tokens stored separately in 1Password",
        "GravityBP","",
        "Team Plan","Team",
        date(2021,1,1),"","Yes",4.00,"Monthly","billing@gravity-bp.com","Card ending 4242",
        "Owner","","","","","","","",
        "github.com","443","HTTPS","",
        "Dev Mukherjee","info@gravity-bp.com",
        "Dev Mukherjee",date(2021,1,1),"Dev Mukherjee",date(2024,6,1),
        date(2024,12,1),date(2024,6,1),"",date(2025,6,1),
        "Dev; GitHub; Code; GBP","3 members. Private repos. PATs tracked separately.","Active",
    ),
]

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 1 — README
# ──────────────────────────────────────────────────────────────────────────────
def create_readme(wb):
    ws = wb.create_sheet("_README")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 85

    write_title(ws, 1, 2, "  Credential Manager — User Guide")

    content = [
        ("PURPOSE", ""),
        ("What is this file?",
         "A centralised credential store for all tenant clients managed by Gravity Business Partners. "
         "Stored on SharePoint; also used as the backend datastore for the Credential Manager application."),
        ("", ""),
        ("SHEETS OVERVIEW", ""),
        ("Master_Credentials",   "Single source of truth. All credentials for all tenants live here."),
        ("Change_Log",           "Audit trail. Every create, update, password change and archive action is logged here."),
        ("Tenants",              "Reference list of all client tenants with contact details."),
        ("Categories",           "Lookup table of all credential categories and subcategories."),
        ("Internal_Users",       "Internal staff who have access to this file."),
        ("Dashboard",            "Live summary: counts by status, category, expiry, and tenant."),
        ("VIEW_*  sheets",       "Read-only live views driven by FILTER() formulas. Data always reflects Master_Credentials."),
        ("", ""),
        ("HOW TO ADD A CREDENTIAL", ""),
        ("Step 1 — Credential ID",  "Use format  CRED-YYYY-NNNN  (e.g. CRED-2025-0016). Increment the last 4-digit number."),
        ("Step 2 — Core fields",    "Fill: Tenant_Code, Category, Service_Name, Username_Email, Password, Status, Priority, Managed_By."),
        ("Step 3 — Save",           "Save the Excel file. The web app will sync automatically on next open."),
        ("Step 4 — Log the action", "Add a row to Change_Log with Action = 'Created' and your name in Changed_By."),
        ("", ""),
        ("HOW TO UPDATE A PASSWORD", ""),
        ("",  "1. Update the Password field in Master_Credentials."),
        ("",  "2. Update Last_Password_Changed to today's date."),
        ("",  "3. Set Last_Updated_By and Last_Updated_Date."),
        ("",  "4. Add a row to Change_Log: Action = 'Password Changed', Field_Changed = 'Password', both value fields = '***'."),
        ("", ""),
        ("STATUS VALUES", ""),
        ("Active",       "Credential is current and working."),
        ("Inactive",     "Temporarily disabled but retained."),
        ("Expired",      "Subscription / access has lapsed. Needs renewal or removal."),
        ("Compromised",  "Suspected or confirmed breach. Change immediately and log."),
        ("Archived",     "No longer needed; kept for historical reference only."),
        ("", ""),
        ("PRIORITY LEVELS", ""),
        ("Critical",     "Business-stopping if lost: global admin, domain registrar, billing accounts."),
        ("High",         "Important service; significant impact if lost."),
        ("Medium",       "Secondary service; moderate impact."),
        ("Low",          "Non-critical; minimal impact."),
        ("", ""),
        ("CREDENTIAL ID FORMAT", ""),
        ("Format",       "CRED-YYYY-NNNN   e.g. CRED-2025-0016"),
        ("Log ID format","LOG-YYYY-NNNN    e.g. LOG-2025-0008"),
        ("Tenant Code",  "3-5 uppercase letters, unique per tenant   e.g. GBP, TAS, BKR"),
        ("", ""),
        ("SECURITY GUIDELINES", ""),
        ("Password field",   "Enter the actual password. This file is protected by SharePoint permissions."),
        ("Sharing",          "Never share or email this file outside the approved SharePoint site."),
        ("API Keys",         "Store full API keys in the API_Key / API_Secret columns. Reference 1Password for extra-sensitive items."),
        ("MFA / 2FA",        "Always record MFA type and backup code location."),
        ("Review Cadence",   "Set Next_Review_Date. Quarterly for Critical/High; bi-annually for Medium/Low."),
        ("", ""),
        ("FILTER VIEWS", ""),
        ("Usage",            "VIEW_* sheets use Excel 365 FILTER() dynamic formulas — they update automatically. "
                             "Do NOT edit data in view sheets; edit in Master_Credentials only."),
        ("Custom filter",    "Use column header dropdowns on Master_Credentials to filter any combination of fields."),
        ("", ""),
        ("FILE INFORMATION", ""),
        ("Generated",        NOW_STR),
        ("Generated by",     "Gravity Business Partners — Credential Manager Setup Script"),
        ("SharePoint path",  "https://<your-tenant>.sharepoint.com/sites/<site>/Shared Documents/Credential_Manager.xlsx"),
        ("Support",          "info@gravity-bp.com"),
    ]

    SECTION_KEYS = {
        "PURPOSE", "SHEETS OVERVIEW", "HOW TO ADD A CREDENTIAL",
        "HOW TO UPDATE A PASSWORD", "STATUS VALUES", "PRIORITY LEVELS",
        "CREDENTIAL ID FORMAT", "SECURITY GUIDELINES", "FILTER VIEWS",
        "FILE INFORMATION",
    }

    for r, (label, value) in enumerate(content, 2):
        ws.row_dimensions[r].height = 17
        a = ws.cell(r, 1, label)
        b = ws.cell(r, 2, value)
        b.alignment = al("left", "center", True)
        if label in SECTION_KEYS:
            a.fill = fl(BLACK); a.font = fn(WHITE, 9, bold=True)
            b.fill = fl(BLACK); b.font = fn(WHITE, 9)
        elif label:
            a.font = fn(TXTCLR, 9, bold=True)
            b.font = fn(TXTCLR, 9)
        else:
            a.font = fn(TXTCLR, 9)
            b.font = fn(TXTCLR, 9)

    ws.sheet_view.showGridLines = False


# ──────────────────────────────────────────────────────────────────────────────
# SHEET 2 — MASTER CREDENTIALS
# ──────────────────────────────────────────────────────────────────────────────
def create_master(wb):
    ws = wb.create_sheet("Master_Credentials")

    for i, (_, w) in enumerate(MASTER_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    write_title(ws, 1, NCOLS_MASTER,
                "  MASTER CREDENTIALS — Single Source of Truth")

    # Row 2 — group labels (dark grey)
    for label, first_name, last_name in MASTER_GROUPS:
        c1 = col_of(first_name)
        c2 = col_of(last_name)
        write_group_header(ws, 2, c1, c2, label)

    # Row 3 — column headers (black)
    ws.row_dimensions[3].height = 30
    for i, (h, _) in enumerate(MASTER_COLS, 1):
        write_header_cell(ws, 3, i, h.replace("_", " "))

    # Data rows
    DATE_COLS = {
        "Subscription_Start", "Subscription_End", "Created_Date",
        "Last_Updated_Date", "Last_Verified_Date", "Last_Password_Changed",
        "Password_Expiry_Date", "Next_Review_Date",
    }
    for row_idx, cred in enumerate(SAMPLE_CREDENTIALS):
        r = row_idx + 4
        ws.row_dimensions[r].height = 17
        for c_idx, (col_name, _) in enumerate(MASTER_COLS, 1):
            val = cred.get(col_name, "")
            fmt = "YYYY-MM-DD" if col_name in DATE_COLS else (
                  "#,##0.00" if col_name == "Monthly_Cost" else None)
            write_data_cell(ws, r, c_idx, val, num_fmt=fmt)

    # Excel Table
    MAX_ROW = 1003
    last_col = get_column_letter(NCOLS_MASTER)
    add_table(ws, "tblCredentials", f"A3:{last_col}{MAX_ROW}", "TableStyleLight1")

    # Data validation
    def col_ltr(name): return get_column_letter(col_of(name))
    def rng(name):     return f"{col_ltr(name)}4:{col_ltr(name)}{MAX_ROW}"

    ws.add_data_validation(dv_list(f'"{",".join(STATUSES)}"',       rng("Status")))
    ws.add_data_validation(dv_list(f'"{",".join(PRIORITIES)}"',     rng("Priority")))
    ws.add_data_validation(dv_list(f'"{",".join(ENVIRONMENTS)}"',   rng("Environment")))
    ws.add_data_validation(dv_list(f'"{",".join(MFA_ENABLED)}"',    rng("MFA_Enabled")))
    ws.add_data_validation(dv_list(f'"{",".join(MFA_TYPES)}"',      rng("MFA_Type")))
    ws.add_data_validation(dv_list(f'"{",".join(ACCESS_LEVELS)}"',  rng("Access_Level")))
    ws.add_data_validation(dv_list(f'"{",".join(PROTOCOLS)}"',      rng("Protocol")))
    ws.add_data_validation(dv_list(f'"{",".join(BILLING_CYCLES)}"', rng("Billing_Cycle")))
    ws.add_data_validation(dv_list(f'"{",".join(AUTO_RENEWAL)}"',   rng("Auto_Renewal")))
    ws.add_data_validation(dv_list(f'"{",".join(RECORD_STATUS)}"',  rng("Record_Status")))

    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False


# ──────────────────────────────────────────────────────────────────────────────
# SHEET 3 — CHANGE LOG
# ──────────────────────────────────────────────────────────────────────────────
def create_changelog(wb):
    ws = wb.create_sheet("Change_Log")
    NCOLS = NCOLS_CHANGELOG

    for i, (_, w) in enumerate(CHANGELOG_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    write_title(ws, 1, NCOLS, "  CHANGE LOG — Full Audit Trail")
    write_info_cell(ws, 2, NCOLS,
        "Log EVERY action here: credential created, updated, password changed, accessed, archived, or compromised.  "
        "Old_Value_Masked and New_Value_Masked should show *** for sensitive fields.")

    ws.row_dimensions[3].height = 28
    for i, (h, _) in enumerate(CHANGELOG_COLS, 1):
        write_header_cell(ws, 3, i, h.replace("_", " "))

    sample_logs = [
        ("LOG-2024-0001", datetime(2024,1,1,9,0),  "CRED-2024-0001","GBP","Gravity Business Partners","Microsoft 365 - GBP Tenant",      "Created",          "All Fields","","",                 "Dev Mukherjee","info@gravity-bp.com","Initial setup",""),
        ("LOG-2024-0002", datetime(2024,3,15,10,30),"CRED-2024-0004","TAS","TechAlpha Solutions",      "Namecheap - techalpha.ae",         "Created",          "All Fields","","",                 "Dev Mukherjee","info@gravity-bp.com","New client onboarding",""),
        ("LOG-2024-0003", datetime(2024,6,1,14,45), "CRED-2024-0001","GBP","Gravity Business Partners","Microsoft 365 - GBP Tenant",      "Password Changed", "Password","***","***",              "Dev Mukherjee","info@gravity-bp.com","Quarterly password rotation",""),
        ("LOG-2024-0004", datetime(2024,6,1,14,50), "CRED-2024-0001","GBP","Gravity Business Partners","Microsoft 365 - GBP Tenant",      "Updated",          "Last_Updated_Date","2024-01-01","2024-06-01","Dev Mukherjee","info@gravity-bp.com","Metadata update after password change",""),
        ("LOG-2024-0005", datetime(2024,7,1,9,0),   "CRED-2024-0007","BKR","Bakery King LLC",          "Google Workspace - bakeryking.com","Updated",          "License_Type","Business Starter","Business Starter","Dev Mukherjee","info@gravity-bp.com","Annual review",""),
        ("LOG-2024-0006", datetime(2024,1,15,11,0), "CRED-2024-0014","MED","MediCare Clinics Group",   "Zoho Mail - medicareclinics.ae",   "Updated",          "Status","Active","Inactive",        "Dev Mukherjee","info@gravity-bp.com","Migrated to M365. Zoho decommissioned.",""),
        ("LOG-2024-0007", datetime(2024,1,15,11,5), "CRED-2024-0009","MED","MediCare Clinics Group",   "Microsoft 365 - MediCare Tenant",  "Created",          "All Fields","","",                 "Dev Mukherjee","info@gravity-bp.com","M365 migration complete",""),
    ]

    for row_idx, row_data in enumerate(sample_logs):
        r = row_idx + 4
        ws.row_dimensions[r].height = 17
        for c_idx, val in enumerate(row_data, 1):
            fmt = "YYYY-MM-DD HH:MM" if c_idx == 2 else None
            write_data_cell(ws, r, c_idx, val, num_fmt=fmt)

    last_col = get_column_letter(NCOLS)
    add_table(ws, "tblChangeLog", f"A3:{last_col}2003", "TableStyleLight1")

    ws.add_data_validation(dv_list(f'"{",".join(LOG_ACTIONS)}"', "G4:G2003"))

    for r in range(4, 2004):
        ws.cell(r, 2).number_format = "YYYY-MM-DD HH:MM"

    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False


# ──────────────────────────────────────────────────────────────────────────────
# SHEET 4 — TENANTS
# ──────────────────────────────────────────────────────────────────────────────
def create_tenants(wb):
    ws = wb.create_sheet("Tenants")
    NCOLS = NCOLS_TENANTS

    for i, (_, w) in enumerate(TENANT_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    write_title(ws, 1, NCOLS, "  TENANTS — Client Reference")
    write_info_cell(ws, 2, NCOLS,
        "Reference table for all clients. Tenant_Code must be unique (3-5 uppercase letters).")

    ws.row_dimensions[3].height = 28
    for i, (h, _) in enumerate(TENANT_COLS, 1):
        write_header_cell(ws, 3, i, h.replace("_", " "))

    for row_idx, row_data in enumerate(TENANT_DATA):
        r = row_idx + 4
        ws.row_dimensions[r].height = 17
        for c_idx, val in enumerate(row_data, 1):
            write_data_cell(ws, r, c_idx, val)

    last_col = get_column_letter(NCOLS)
    add_table(ws, "tblTenants", f"A3:{last_col}103", "TableStyleLight1")
    ws.add_data_validation(dv_list(f'"{",".join(TENANT_STATUSES)}"', "K4:K103"))

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


# ──────────────────────────────────────────────────────────────────────────────
# SHEET 5 — CATEGORIES
# ──────────────────────────────────────────────────────────────────────────────
def create_categories(wb):
    ws = wb.create_sheet("Categories")
    NCOLS = NCOLS_CATEGORIES

    for i, (_, w) in enumerate(CATEGORY_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    write_title(ws, 1, NCOLS, "  CATEGORIES — Credential Category Reference")
    write_info_cell(ws, 2, NCOLS,
        "Lookup reference for categories used in Master_Credentials. Do not delete rows.")

    ws.row_dimensions[3].height = 28
    for i, (h, _) in enumerate(CATEGORY_COLS, 1):
        write_header_cell(ws, 3, i, h.replace("_", " "))

    for row_idx, row_data in enumerate(CATEGORIES_DATA):
        r = row_idx + 4
        ws.row_dimensions[r].height = 22
        for c_idx, val in enumerate(row_data, 1):
            write_data_cell(ws, r, c_idx, val)

    last_col = get_column_letter(NCOLS)
    add_table(ws, "tblCategories", f"A3:{last_col}53", "TableStyleLight1")

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


# ──────────────────────────────────────────────────────────────────────────────
# SHEET 6 — INTERNAL USERS
# ──────────────────────────────────────────────────────────────────────────────
def create_users(wb):
    ws = wb.create_sheet("Internal_Users")
    NCOLS = NCOLS_USERS

    for i, (_, w) in enumerate(USER_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    write_title(ws, 1, NCOLS, "  INTERNAL USERS — Staff Access List")
    write_info_cell(ws, 2, NCOLS,
        "Staff members with access to this credential store. Remove leavers immediately.")

    ws.row_dimensions[3].height = 28
    for i, (h, _) in enumerate(USER_COLS, 1):
        write_header_cell(ws, 3, i, h.replace("_", " "))

    for row_idx, row_data in enumerate(USER_DATA):
        r = row_idx + 4
        ws.row_dimensions[r].height = 17
        for c_idx, val in enumerate(row_data, 1):
            write_data_cell(ws, r, c_idx, val)

    last_col = get_column_letter(NCOLS)
    add_table(ws, "tblUsers", f"A3:{last_col}53", "TableStyleLight1")

    ws.add_data_validation(dv_list(f'"{",".join(USER_ROLES)}"',   "D4:D53"))
    ws.add_data_validation(dv_list(f'"{",".join(USER_ACCESS)}"',  "F4:F53"))
    ws.add_data_validation(dv_list(f'"{",".join(USER_STATUSES)}"',"G4:G53"))

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


# ──────────────────────────────────────────────────────────────────────────────
# SHEET 7 — DASHBOARD
# ──────────────────────────────────────────────────────────────────────────────
def create_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    for col, w in zip("ABCDE", [32, 14, 32, 14, 4]):
        ws.column_dimensions[col].width = w

    write_title(ws, 1, 4, "  DASHBOARD — Live Summary")

    def section(row, title):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row, 1, f"  {title}")
        c.fill = fl(DGREY); c.font = fn(WHITE, 9, bold=True)
        c.alignment = al("left", "center")
        ws.row_dimensions[row].height = 20

    def hdr(row, *labels):
        ws.row_dimensions[row].height = 18
        for col, label in enumerate(labels, 1):
            c = ws.cell(row, col, label)
            c.fill = fl("E8E8E8"); c.font = fn(TXTCLR, 9, bold=True)
            c.alignment = al("center", "center")
            c.border = thin_border()

    def metric_row(row, label1, formula1, label2=None, formula2=None):
        ws.row_dimensions[row].height = 20
        for col, label, formula in [(1, label1, formula1), (3, label2, formula2)]:
            if label is None:
                continue
            lc = ws.cell(row, col, label)
            lc.font = fn(TXTCLR, 9); lc.alignment = al("left", "center")
            lc.border = thin_border()
            vc = ws.cell(row, col + 1, formula)
            vc.font = fn(TXTCLR, 11, bold=True); vc.alignment = al("center", "center")
            vc.border = thin_border()

    # Status
    section(3, "STATUS BREAKDOWN")
    hdr(4, "Status", "Count", "Status", "Count")
    metric_row(5,  "Active",      '=COUNTIF(tblCredentials[Status],"Active")',
                   "Inactive",   '=COUNTIF(tblCredentials[Status],"Inactive")')
    metric_row(6,  "Expired",     '=COUNTIF(tblCredentials[Status],"Expired")',
                   "Compromised",'=COUNTIF(tblCredentials[Status],"Compromised")')
    metric_row(7,  "Archived",    '=COUNTIF(tblCredentials[Status],"Archived")',
                   "TOTAL",      '=COUNTA(tblCredentials[CredentialID])')

    # Priority
    section(9, "PRIORITY BREAKDOWN")
    hdr(10, "Priority", "Count", "Priority", "Count")
    metric_row(11, "Critical", '=COUNTIF(tblCredentials[Priority],"Critical")',
                   "High",     '=COUNTIF(tblCredentials[Priority],"High")')
    metric_row(12, "Medium",   '=COUNTIF(tblCredentials[Priority],"Medium")',
                   "Low",      '=COUNTIF(tblCredentials[Priority],"Low")')

    # Expiry alerts
    section(14, "EXPIRY ALERTS")
    hdr(15, "Alert", "Count")
    expiry_items = [
        (16, "Expiring within 30 days",
             '=COUNTIFS(tblCredentials[Subscription_End],">="&TODAY(),tblCredentials[Subscription_End],"<="&(TODAY()+30))'),
        (17, "Expiring within 90 days",
             '=COUNTIFS(tblCredentials[Subscription_End],">="&TODAY(),tblCredentials[Subscription_End],"<="&(TODAY()+90))'),
        (18, "Already expired",
             '=COUNTIFS(tblCredentials[Subscription_End],"<"&TODAY(),tblCredentials[Subscription_End],"<>")'),
        (19, "MFA not enabled",
             '=COUNTIF(tblCredentials[MFA_Enabled],"No")'),
    ]
    for r, label, formula in expiry_items:
        ws.row_dimensions[r].height = 20
        lc = ws.cell(r, 1, label); lc.font = fn(TXTCLR,9); lc.alignment = al("left","center"); lc.border = thin_border()
        vc = ws.cell(r, 2, formula); vc.font = fn(TXTCLR,11,bold=True); vc.alignment = al("center","center"); vc.border = thin_border()

    # Category breakdown
    section(21, "CREDENTIALS BY CATEGORY")
    hdr(22, "Category", "Count", "Category", "Count")
    cats = [row[1] for row in CATEGORIES_DATA]
    mid  = len(cats) // 2 + len(cats) % 2
    left_cats, right_cats = cats[:mid], cats[mid:]
    for i, cat in enumerate(left_cats):
        r = 23 + i
        ws.row_dimensions[r].height = 18
        lc = ws.cell(r,1,cat); lc.font=fn(TXTCLR,9); lc.alignment=al("left","center"); lc.border=thin_border()
        vc = ws.cell(r,2,f'=COUNTIF(tblCredentials[Category],"{cat}")'); vc.font=fn(TXTCLR,10,bold=True); vc.alignment=al("center","center"); vc.border=thin_border()
    for i, cat in enumerate(right_cats):
        r = 23 + i
        lc = ws.cell(r,3,cat); lc.font=fn(TXTCLR,9); lc.alignment=al("left","center"); lc.border=thin_border()
        vc = ws.cell(r,4,f'=COUNTIF(tblCredentials[Category],"{cat}")'); vc.font=fn(TXTCLR,10,bold=True); vc.alignment=al("center","center"); vc.border=thin_border()

    # Tenant breakdown
    tr_start = 23 + len(left_cats) + 2
    section(tr_start, "CREDENTIALS BY TENANT")
    hdr(tr_start+1, "Tenant", "Code", "Count")
    for i, (tid, code, name, *_) in enumerate(TENANT_DATA):
        r = tr_start + 2 + i
        ws.row_dimensions[r].height = 18
        lc = ws.cell(r,1,name);  lc.font=fn(TXTCLR,9); lc.alignment=al("left","center");   lc.border=thin_border()
        cc = ws.cell(r,2,code);  cc.font=fn(TXTCLR,9,bold=True); cc.alignment=al("center","center"); cc.border=thin_border()
        vc = ws.cell(r,3,f'=COUNTIF(tblCredentials[Tenant_Code],"{code}")'); vc.font=fn(TXTCLR,10,bold=True); vc.alignment=al("center","center"); vc.border=thin_border()

    # Refresh stamp
    last_r = tr_start + 2 + len(TENANT_DATA) + 2
    ws.merge_cells(start_row=last_r, start_column=1, end_row=last_r, end_column=4)
    stamp = ws.cell(last_r, 1, '=TODAY()')
    stamp.number_format = '"Refreshed: "YYYY-MM-DD'
    stamp.font = fn("888888", 8, italic=True)
    stamp.alignment = al("center", "center")


# ──────────────────────────────────────────────────────────────────────────────
# FILTER VIEW SHEETS
# ──────────────────────────────────────────────────────────────────────────────
def create_filter_view(wb, sheet_name, title_text, filter_formula, info_text):
    ws = wb.create_sheet(sheet_name)

    for i, (_, w) in enumerate(MASTER_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    write_title(ws, 1, NCOLS_MASTER, f"  {title_text}")
    write_info_cell(ws, 2, NCOLS_MASTER,
        f"LIVE VIEW  |  {info_text}  |  "
        "Updates automatically from Master_Credentials. Edit data in Master_Credentials only.")

    ws.row_dimensions[3].height = 28
    for i, (h, _) in enumerate(MASTER_COLS, 1):
        write_header_cell(ws, 3, i, h.replace("_", " "))

    ws.cell(4, 1).value = filter_formula
    ws.cell(4, 1).font  = fn(TXTCLR, 9)

    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────
def main():
    print("Building Credential Manager workbook ...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("  [1/13]  _README")
    create_readme(wb)

    print("  [2/13]  Master_Credentials")
    create_master(wb)

    print("  [3/13]  Change_Log")
    create_changelog(wb)

    print("  [4/13]  Tenants")
    create_tenants(wb)

    print("  [5/13]  Categories")
    create_categories(wb)

    print("  [6/13]  Internal_Users")
    create_users(wb)

    print("  [7/13]  Dashboard")
    create_dashboard(wb)

    views = [
        ("VIEW_M365_Admin",
         "VIEW — Microsoft 365 Credentials",
         '=IFERROR(FILTER(tblCredentials,tblCredentials[Category]="Microsoft 365","No Microsoft 365 credentials found."),"Error loading data.")',
         "Microsoft 365 admin and generic accounts"),

        ("VIEW_Cloud_Infra",
         "VIEW — Cloud Infrastructure Credentials",
         '=IFERROR(FILTER(tblCredentials,tblCredentials[Category]="Cloud Infrastructure","No Cloud Infrastructure credentials found."),"Error loading data.")',
         "AWS | Azure | DigitalOcean portals"),

        ("VIEW_Domains_FTP",
         "VIEW — Domain, Hosting & FTP Credentials",
         '=IFERROR(FILTER(tblCredentials,(tblCredentials[Category]="Domain & DNS")+(tblCredentials[Category]="Web Hosting & Panels")+(tblCredentials[Category]="FTP & File Sharing"),"No records found."),"Error loading data.")',
         "Domain & DNS | Web Hosting & Panels | FTP & File Sharing"),

        ("VIEW_Social_Media",
         "VIEW — Social Media Credentials",
         '=IFERROR(FILTER(tblCredentials,tblCredentials[Category]="Social Media","No Social Media credentials found."),"Error loading data.")',
         "Social Media handles and brand channels"),

        ("VIEW_Marketing_Creative",
         "VIEW — Marketing & Creative Credentials",
         '=IFERROR(FILTER(tblCredentials,tblCredentials[Category]="Marketing & Creative","No Marketing & Creative credentials found."),"Error loading data.")',
         "Marketing tools, design, and creative suite logins"),

        ("VIEW_ERP_Finance",
         "VIEW — ERP & Finance Credentials",
         '=IFERROR(FILTER(tblCredentials,(tblCredentials[Category]="ERP & Business Portals")+(tblCredentials[Category]="Finance & Accounting"),"No ERP & Finance credentials found."),"Error loading data.")',
         "Zoho | Odoo | Accounting | Invoicing | Payments"),

        ("VIEW_Expiring_90d",
         "VIEW — Credentials Expiring Within 90 Days",
         '=IFERROR(FILTER(tblCredentials,(tblCredentials[Subscription_End]-TODAY()<=90)*(tblCredentials[Subscription_End]>=TODAY()),"No credentials expiring within 90 days."),"Error loading data.")',
         "Subscription_End is within the next 90 days"),
    ]

    for idx, (name, title, formula, info) in enumerate(views, 8):
        print(f"  [{idx}/13]  {name}")
        create_filter_view(wb, name, title, formula, info)

    print(f"\n  Saving to: {OUTPUT_FILE} ...")
    wb.save(OUTPUT_FILE)
    print(f"\nDone!  Workbook saved to:\n    {OUTPUT_FILE}\n")
    print("Sheets created:")
    for ws in wb.worksheets:
        print(f"  - {ws.title}")


if __name__ == "__main__":
    main()
