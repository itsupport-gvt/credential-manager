#!/usr/bin/env python3
"""
tools/generate_excel_template.py
Generate the Credential_Manager.xlsx template file.

Creates five auto-growing Excel Tables (no test data), pre-populated with the
predefined category list and dropdown validation on key columns.

Usage:
    python tools/generate_excel_template.py [output_path]

Defaults to Credential_Manager.xlsx in the current directory.
Requires: pip install openpyxl
"""

from __future__ import annotations

import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("ERROR: openpyxl is required.\n  pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Palette – matches the app's design tokens
# ---------------------------------------------------------------------------

C_HEADER_BG  = "111111"   # pure dark for headers
C_HEADER_FG  = "FFFFFF"   # pure white text
C_ALT_ROW    = "F9F9F9"   # very subtle light grey for alternating rows
C_BORDER     = "EAEAEA"   # minimal light border
C_SECTION_BG = "F4F4F4"   # section header bg


# ---------------------------------------------------------------------------
# Column definitions  (name, col_width, comment/hint)
# ---------------------------------------------------------------------------

CRED_COLS: list[tuple[str, int, str]] = [
    ("CredentialID",          18, "Auto-generated. Format: CRED-YYYY-NNNN"),
    ("Credential_Type",       16, "Password | OTP-Only | API Key | OAuth2 | Database | SSH | License Key | Certificate | Custom"),
    ("Tenant_Code",           14, "Short code identifying the client, e.g. GVT"),
    ("Tenant_Name",           22, "Full name of the tenant / client"),
    ("Category",              20, "Must match a Category_Name in tblCategories"),
    ("Subcategory",           20, "Optional sub-classification"),
    ("Service_Name",          26, "Human-readable name, e.g. Microsoft 365 Admin"),
    ("Service_URL",           32, "https://…"),
    ("Environment",           14, "Production | Staging | Development | Testing | DR"),
    ("Status",                12, "Active | Inactive | Expired | Compromised | Archived"),
    ("Priority",              12, "Critical | High | Medium | Low"),
    ("Username_Email",        28, "Login username or email address"),
    ("Password",              28, "Plain text — encrypted on import into the app"),
    ("Recovery_Email",        26, ""),
    ("Recovery_Phone",        18, ""),
    ("MFA_Enabled",           12, "Yes | No"),
    ("MFA_Type",              16, "TOTP | SMS | Email | Hardware Key | Passkey | Other"),
    ("MFA_App_Name",          22, "e.g. Microsoft Authenticator, Google Authenticator"),
    ("Backup_Codes_Location", 24, "Where the backup codes are stored"),
    ("Security_Notes",        32, "Additional security information"),
    ("Account_Display_Name",  26, "Display name shown on the account"),
    ("Account_ID",            22, ""),
    ("License_Type",          18, ""),
    ("Plan_Tier",             16, "e.g. Business Premium, Enterprise"),
    ("Subscription_Start",    18, "YYYY-MM-DD"),
    ("Subscription_End",      18, "YYYY-MM-DD"),
    ("Auto_Renewal",          14, "Yes | No | Unknown"),
    ("Monthly_Cost",          14, "Numeric (USD)"),
    ("Billing_Cycle",         16, "Monthly | Annual | Quarterly | Bi-Annual | One-Time"),
    ("Billing_Email",         26, ""),
    ("Payment_Reference",     22, "Invoice or payment ref number"),
    ("Access_Level",          18, "Admin | Owner | Member | Viewer | Read-Only | Service Account"),
    ("Linked_Credential_ID",  22, "CredentialID this is linked to, if any"),
    ("API_Key",               32, "Plain text — encrypted on import"),
    ("API_Secret",            32, "Plain text — encrypted on import"),
    ("Client_ID",             34, "OAuth2 / app Client ID"),
    ("Client_Secret",         34, "Plain text — encrypted on import"),
    ("Tenant_ID_App",         36, "Azure AD / application tenant ID"),
    ("Subscription_ID_Azure", 36, "Azure Subscription ID"),
    ("Server_Hostname",       26, "Hostname or IP address"),
    ("Port",                  10, ""),
    ("Protocol",              14, "HTTPS | HTTP | SFTP | FTP | SSH | RDP | MySQL | PostgreSQL | MSSQL | Other"),
    ("Database_Name",         22, ""),
    ("Managed_By",            22, "Name of person responsible"),
    ("Managed_By_Email",      28, ""),
    ("Created_By",            22, ""),
    ("Created_Date",          14, "YYYY-MM-DD"),
    ("Last_Updated_By",       22, ""),
    ("Last_Updated_Date",     14, "YYYY-MM-DD"),
    ("Last_Verified_Date",    18, "YYYY-MM-DD"),
    ("Last_Password_Changed", 20, "YYYY-MM-DD"),
    ("Password_Expiry_Date",  20, "YYYY-MM-DD — leave blank if no expiry"),
    ("Next_Review_Date",      18, "YYYY-MM-DD"),
    ("Tags",                  26, "Comma-separated keywords for filtering"),
    ("Notes",                 44, "Free-text notes"),
    ("Record_Status",         14, "Active | Archived"),
]

LOG_COLS: list[tuple[str, int, str]] = [
    ("LogID",            36, "UUID — auto-generated by the app"),
    ("Timestamp",        24, "ISO 8601 datetime"),
    ("CredentialID",     18, ""),
    ("Tenant_Code",      14, ""),
    ("Tenant_Name",      22, ""),
    ("Service_Name",     26, ""),
    ("Action",           14, "CREATE | UPDATE | DELETE | ARCHIVE | REVEAL | ACCESS"),
    ("Field_Changed",    26, "Which field was modified"),
    ("Old_Value_Masked", 26, "Previous value (masked for secrets)"),
    ("New_Value_Masked", 26, "New value (masked for secrets)"),
    ("Changed_By",       22, ""),
    ("Changed_By_Email", 28, ""),
    ("Reason",           32, ""),
    ("Notes",            44, ""),
]

TENANT_COLS: list[tuple[str, int, str]] = [
    ("TenantID",        36, "UUID — auto-generated"),
    ("Tenant_Code",     14, "Unique short code, e.g. GVT"),
    ("Tenant_Name",     26, "Full legal or trading name"),
    ("Industry",        22, "e.g. Financial Services, Healthcare"),
    ("Primary_Contact", 26, ""),
    ("Contact_Email",   30, ""),
    ("Contact_Phone",   18, ""),
    ("Account_Manager", 24, "Internal account owner"),
    ("Contract_Start",  16, "YYYY-MM-DD"),
    ("Contract_End",    16, "YYYY-MM-DD"),
    ("Status",          12, "Active | Inactive"),
    ("Notes",           44, ""),
]

CAT_COLS: list[tuple[str, int, str]] = [
    ("CategoryID",    16, "Unique ID, e.g. CAT-CLOUD"),
    ("Category_Name", 30, "Display name"),
    ("Category_Code", 16, "Short code, e.g. CLOUD"),
    ("Description",   44, ""),
    ("Subcategories", 64, "Semicolon-separated: Sub1; Sub2; Sub3"),
]

USER_COLS: list[tuple[str, int, str]] = [
    ("UserID",       36, "UUID — auto-generated"),
    ("Full_Name",    28, ""),
    ("Email",        32, ""),
    ("Role",         22, "e.g. IT Admin, MSP Engineer"),
    ("Department",   22, ""),
    ("Access_Level", 16, "Admin | Standard | ReadOnly"),
    ("Status",       12, "Active | Inactive"),
    ("Notes",        44, ""),
]


# ---------------------------------------------------------------------------
# Pre-populated category data
# ---------------------------------------------------------------------------

CATEGORIES: list[tuple] = [
    ("CAT-M365", "Microsoft 365", "M365", "Microsoft 365 tenant and service accounts",
     "Global Admin;Exchange Admin;SharePoint Admin;Teams Admin;User Account;Mailbox Access (OTP/SSO);Service Account"),
    ("CAT-CLOUD", "Cloud Infrastructure", "CLOUD", "Amazon Web Services, Microsoft Azure, DigitalOcean and other cloud providers",
     "AWS Root;AWS IAM User;Azure Portal;Azure Sub Owner;DigitalOcean;Google Cloud Portal;Service Principal"),
    ("CAT-DNS", "Domain & DNS", "DNS", "Domain registrars, DNS hosting and SSL certificates",
     "Domain Registrar;DNS Provider;SSL/TLS Certificate;CDN;Domain Forwarding"),
    ("CAT-HOST", "Web Hosting & Panels", "HOST", "Server control panels, cPanel, WHM, Plesk, and hosting portal access",
     "cPanel;Plesk;WHM;Web Host Portal;Virtual Private Server (VPS);Managed Hosting"),
    ("CAT-CMS", "CMS & Websites", "CMS", "Content Management Systems (WordPress, Django, CMS portals) and site admin panels",
     "WordPress Admin;Django Admin;Shopify Admin;Webflow;Squarespace;Custom CMS"),
    ("CAT-GOOG", "Google Ecosystem", "GOOG", "Google Search Console, Google Business Profile, Analytics and Google account credentials",
     "Google Search Console;Google Business Profile (GMB);Google Analytics;Google Tag Manager;Workspace Admin"),
    ("CAT-DB", "Databases", "DB", "Database administration, SQL Server, Azure SQL, MySQL, and PostgreSQL logins",
     "Azure SQL;MySQL/MariaDB;PostgreSQL;Microsoft SQL Server;MongoDB;Redis;Database Administrator"),
    ("CAT-FTP", "FTP & File Sharing", "FTP", "File transfer protocol (FTP, SFTP) credentials and cloud/network storage vaults",
     "SFTP;FTP;FTPS;S3 Bucket;WebDAV;Network File Share"),
    ("CAT-ERP", "ERP & Business Portals", "ERP", "Zoho, Odoo, task workflow portals, and learning platforms",
     "Zoho Admin;Zoho User;Odoo Admin;Odoo User;ERP Portal;Task Assignment;Workflow Portal"),
    ("CAT-MKT", "Marketing & Creative", "MKT", "Marketing tools, graphic design, and creative suite applications",
     "Adobe Creative Cloud;Graphic Design Tool;Email Marketing (Mailchimp);SEO/SEM Tool;Marketing Analytics;Canva"),
    ("CAT-SOC", "Social Media", "SOC", "Brand social media handles and ad manager accounts",
     "Facebook/Meta;Instagram;Twitter/X;LinkedIn;YouTube;TikTok;Pinterest"),
    ("CAT-AI", "AI & Chatbots", "AI", "Artificial Intelligence systems, LLM tools, and AI developer portals",
     "ChatGPT (OpenAI);Claude (Anthropic);Gemini (Google);Copilot;AI API Key;Custom Chatbot"),
    ("CAT-HR", "HR & Recruitment", "HR", "HR management systems (HRMS), job portals (Naukri, Indeed), and payroll portal access",
     "HRMS Portal;Recruitment Portal (Naukri/Indeed);Payroll System;Attendance/Leave System"),
    ("CAT-FIN", "Finance & Accounting", "FIN", "Accounting softwares (Zoho Admin, Odoo Admin, users, etc.) and invoicing platforms",
     "Zoho Books/Finance;Odoo Accounting;Accounting Software;Payment Gateway;Billing/Invoice Portal"),
    ("CAT-WORK", "Workflow & Productivity", "WORK", "Team collaboration, task tracking, and learning portals",
     "Jira;Monday.com;Asana;Trello;ClickUp;Slack;Microsoft Teams;Learning Portal (Udemy)"),
    ("CAT-NET", "Network & CCTV", "NET", "Router/switch logins, internal WiFi, and CCTV/NVR surveillance cameras",
     "WiFi Router/Access Point;Firewall/Switch;CCTV Camera/NVR;Network Attached Storage (NAS);ISP Admin Portal"),
    ("CAT-TEL", "Telecom & Utility", "TEL", "Network service providers, postpaid sim portals, and broadband logins",
     "Postpaid SIM Portal;WiFi Service Provider;Telecom Login;Broadband Portal"),
    ("CAT-OTH", "Other", "OTH", "Miscellaneous credentials not covered in above categories",
     "General;Utility;Hardware/IoT;Security Vault"),
]


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, size: int = 10, color: str = "000000") -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _thin_border() -> Border:
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _left(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


# ---------------------------------------------------------------------------
# Core sheet builder
# ---------------------------------------------------------------------------

def _build_sheet(
    wb: Workbook,
    sheet_name: str,
    columns: list[tuple[str, int, str]],
    table_name: str,
    data_rows: list[tuple] | None = None,
) -> None:
    ws = wb.create_sheet(sheet_name)

    hdr_fill  = _fill(C_HEADER_BG)
    hdr_font  = _font(bold=True, size=10, color=C_HEADER_FG)
    alt_fill  = _fill(C_ALT_ROW)
    data_font = _font(size=10)

    # ── Column headers ───────────────────────────────────────────────────────
    for ci, (col_name, col_width, comment) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = _center()
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    ws.row_dimensions[1].height = 26

    # ── Data rows ────────────────────────────────────────────────────────────
    n_data = 0
    if data_rows:
        for ri, row_vals in enumerate(data_rows, start=2):
            fill = alt_fill if ri % 2 == 0 else None
            for ci, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = data_font
                cell.alignment = _left()
                if fill:
                    cell.fill = fill
            ws.row_dimensions[ri].height = 18
        n_data = len(data_rows)

    # Table needs at least header + 1 row; use an empty placeholder row if no data.
    table_last_row = max(2, 1 + n_data)

    # ── Excel Table ──────────────────────────────────────────────────────────
    last_col_letter = get_column_letter(len(columns))
    table_ref       = f"A1:{last_col_letter}{table_last_row}"

    tbl = Table(displayName=table_name, ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)

    # ── Freeze header row ────────────────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True


# ---------------------------------------------------------------------------
# Data-validation helpers
# ---------------------------------------------------------------------------

def _dv_list(formula: str, sq_ref: str) -> DataValidation:
    """Create a list data-validation for *sq_ref* (e.g. 'B2:B1048576')."""
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=False)
    dv.sqref = sq_ref
    return dv


def _col_ref(ws_cols: list[tuple], col_name: str, start_row: int = 2) -> str | None:
    """Return 'XN:X1048576' for *col_name* in *ws_cols*, or None if not found."""
    for i, (name, _, _) in enumerate(ws_cols, start=1):
        if name == col_name:
            letter = get_column_letter(i)
            return f"{letter}{start_row}:{letter}1048576"
    return None


# ---------------------------------------------------------------------------
# Cover sheet
# ---------------------------------------------------------------------------

def _build_cover(wb: Workbook) -> None:
    ws = wb.create_sheet("About", 0)
    ws.sheet_view.showGridLines = False

    hdr_fill = _fill(C_HEADER_BG)
    blue_fill = _fill("F4F4F4")

    def row(r: int, text: str, bold: bool = False, size: int = 11, color: str = "111111", bg: str | None = None) -> None:
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = Font(name="Calibri", bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if bg:
            cell.fill = _fill(bg)
        ws.row_dimensions[r].height = 20 if not bold else 26

    ws.column_dimensions["A"].width = 90
    ws.merge_cells("A1:A1")
    title_cell = ws.cell(row=1, column=1, value="Credential Manager  ·  Gravity Business Partners")
    title_cell.fill = hdr_fill
    title_cell.font = Font(name="Calibri", bold=True, size=16, color=C_HEADER_FG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40

    row(2, "")
    row(3, "This workbook is the SharePoint backend for the Credential Manager desktop app.",
        size=11, color="3C4043")
    row(4, "Data is synced automatically between the app and this file via the Microsoft Graph API.",
        size=10, color="5F6368")
    row(5, "")
    row(6, "SHEETS", bold=True, size=12, color=C_HEADER_BG)
    items = [
        ("Credentials",   "All credential records — the primary table (tblCredentials)"),
        ("ChangeLog",     "Audit trail of every create / update / delete / reveal action (tblChangeLog)"),
        ("Tenants",       "Client / tenant directory (tblTenants)"),
        ("Categories",    "Credential categories and subcategories (tblCategories)"),
        ("Users",         "User directory — who has access to the app (tblUsers)"),
    ]
    for i, (sheet, desc) in enumerate(items, start=7):
        cell = ws.cell(row=i, column=1, value=f"  {sheet:16s}  {desc}")
        cell.font = Font(name="Calibri", size=10, color="111111")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if i % 2 == 0:
            cell.fill = _fill("F9F9F9")
        ws.row_dimensions[i].height = 20

    row(12, "")
    row(13, "IMPORTANT NOTES", bold=True, size=12, color=C_HEADER_BG)
    notes = [
        "• Never delete column headers — the sync service relies on exact header names.",
        "• Do not rename table names (tblCredentials, tblChangeLog, etc.).",
        "• Password, API_Key, API_Secret and Client_Secret are stored as plain text here but",
        "  are encrypted in the app's local database. Protect this file accordingly.",
        "• The tables auto-grow: type data in the row immediately below the last row.",
        "• Date columns must use YYYY-MM-DD format for reliable sync.",
    ]
    for i, note in enumerate(notes, start=14):
        cell = ws.cell(row=i, column=1, value=note)
        cell.font = Font(name="Calibri", size=10, color="333333")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 20

    ws.row_dimensions[20].height = 30


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def generate(output_path: Path) -> None:
    wb = Workbook()
    # Remove the default blank sheet
    wb.remove(wb.active)

    # ── Cover ────────────────────────────────────────────────────────────────
    _build_cover(wb)

    # ── Credentials ──────────────────────────────────────────────────────────
    _build_sheet(wb, "Credentials", CRED_COLS, "tblCredentials")
    ws_cred = wb["Credentials"]

    # Data validation for Credentials columns
    _dv_defs_cred = [
        ("Credential_Type", '"Password,OTP-Only,API Key,OAuth2,Database,SSH,License Key,Certificate,Custom"'),
        ("Environment",     '"Production,Staging,Development,Testing,DR"'),
        ("Status",          '"Active,Inactive,Expired,Compromised,Archived"'),
        ("Priority",        '"Critical,High,Medium,Low"'),
        ("MFA_Enabled",     '"Yes,No"'),
        ("MFA_Type",        '"TOTP,SMS,Email,Hardware Key,Passkey,Push,Biometric,Other"'),
        ("Auto_Renewal",    '"Yes,No,Unknown"'),
        ("Billing_Cycle",   '"Monthly,Annual,Quarterly,Bi-Annual,One-Time"'),
        ("Access_Level",    '"Admin,Owner,Member,Viewer,Read-Only,Service Account"'),
        ("Protocol",        '"HTTPS,HTTP,SFTP,FTP,SSH,RDP,MySQL,PostgreSQL,MSSQL,Other"'),
        ("Record_Status",   '"Active,Archived"'),
    ]
    for col_name, formula in _dv_defs_cred:
        sq = _col_ref(CRED_COLS, col_name)
        if sq:
            ws_cred.add_data_validation(_dv_list(formula, sq))

    # ── ChangeLog ────────────────────────────────────────────────────────────
    _build_sheet(wb, "ChangeLog", LOG_COLS, "tblChangeLog")
    ws_log = wb["ChangeLog"]
    sq_action = _col_ref(LOG_COLS, "Action")
    if sq_action:
        ws_log.add_data_validation(_dv_list('"CREATE,UPDATE,DELETE,ARCHIVE,REVEAL,ACCESS"', sq_action))

    # ── Tenants ──────────────────────────────────────────────────────────────
    _build_sheet(wb, "Tenants", TENANT_COLS, "tblTenants")
    ws_tenant = wb["Tenants"]
    sq_t_status = _col_ref(TENANT_COLS, "Status")
    if sq_t_status:
        ws_tenant.add_data_validation(_dv_list('"Active,Inactive"', sq_t_status))

    # ── Categories  (pre-populated) ──────────────────────────────────────────
    _build_sheet(wb, "Categories", CAT_COLS, "tblCategories", data_rows=CATEGORIES)

    # ── Users ────────────────────────────────────────────────────────────────
    _build_sheet(wb, "Users", USER_COLS, "tblUsers")
    ws_users = wb["Users"]
    sq_u_access = _col_ref(USER_COLS, "Access_Level")
    sq_u_status = _col_ref(USER_COLS, "Status")
    if sq_u_access:
        ws_users.add_data_validation(_dv_list('"Admin,Standard,ReadOnly"', sq_u_access))
    if sq_u_status:
        ws_users.add_data_validation(_dv_list('"Active,Inactive"', sq_u_status))

    # ── Tab colours ──────────────────────────────────────────────────────────
    tab_colors = {
        "About":       "000000",
        "Credentials": "222222",
        "ChangeLog":   "444444",
        "Tenants":     "666666",
        "Categories":  "888888",
        "Users":       "AAAAAA",
    }
    for sheet_name, color in tab_colors.items():
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_properties.tabColor = color

    # ── Save ─────────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"Template saved: {output_path}")
    print(f"  Sheets   : {', '.join(wb.sheetnames)}")
    print(f"  Cred cols: {len(CRED_COLS)}")
    print(f"  Categories pre-loaded: {len(CATEGORIES)}")


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Credential_Manager.xlsx")
    generate(out)
